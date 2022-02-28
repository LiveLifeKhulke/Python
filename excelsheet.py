import openpyxl
book=openpyxl.load_workbook("C:\\Users\\degahlot\\Desktop\\Book1.xlsx")
sheet = book.active
cell=sheet['A3'].value
print(cell)
Dict={}
sheet.cell(row=2,column=2).value= "Deepak"
print(sheet.cell(row=2,column=2).value)

for i in range(1,sheet.max_row+1):
    if sheet.cell(row=i,column=1).value=='Case2':
        for j in range(1,sheet.max_column+1):
            Dict[sheet.cell(row=1,column=j).value]=sheet.cell(row=i,column=j).value

print(Dict)